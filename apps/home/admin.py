from django.contrib import admin, messages
from django.shortcuts import render, redirect
from django.urls import path
from django import forms
import openpyxl

from .models import Visit, Invoice, InvoiceItem, Product, AgedReceivable

# Inline for Invoice Items
class InvoiceItemInline(admin.TabularInline):
    model = InvoiceItem
    extra = 0
    fields = ('product', 'quantity', 'brand', 'part_number')
    show_change_link = False

# Invoice Admin with Items
@admin.register(Invoice)
class InvoiceAdmin(admin.ModelAdmin):
    list_display = (
        'invoice_number', 'client', 'invoice_date',
        'due_date', 'salesperson', 'subtotal', 'total',
    )
    inlines = [InvoiceItemInline]
    list_filter = ('invoice_date', 'due_date', 'salesperson', 'client')
    search_fields = ('invoice_number', 'client', 'salesperson')
    ordering = ('-invoice_date',)

# Excel Upload Form (reused)
class ExcelUploadForm(forms.Form):
    excel_file = forms.FileField(label="Upload Excel (.xlsx)")

# Product Admin with Excel upload
@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('part_number', 'name', 'brand', 'quantity_on_hand', 'unit_of_measure', 'sales_price')
    change_list_template = "admin/product_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel))
        ]
        return custom_urls + urls

    def upload_excel(self, request):
        if request.method == "POST":
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = form.cleaned_data["excel_file"]
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                headers = [cell.value.strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                header_map = {header: idx for idx, header in enumerate(headers)}

                required_headers = ['part number', 'name', 'brand', 'quantity', 'uom', 'price']
                if not all(key in header_map for key in required_headers):
                    messages.error(request, "Missing required columns in Excel.")
                    return redirect("..")

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        Product.objects.update_or_create(
                            part_number=row[header_map['part number']],
                            defaults={
                                'name': row[header_map['name']],
                                'brand': row[header_map['brand']],
                                'quantity_on_hand': int(row[header_map['quantity']]),
                                'unit_of_measure': row[header_map['uom']],
                                'sales_price': float(str(row[header_map['price']]).replace(",", ""))
                            }
                        )
                    except Exception as e:
                        messages.warning(request, f"Row error: {e}")

                messages.success(request, "✅ Products uploaded successfully!")
                return redirect("..")
        else:
            form = ExcelUploadForm()

        return render(request, "admin/product_upload_form.html", {
            'form': form,
            'title': "Upload Product Excel"
        })

# Aged Receivable Admin
@admin.register(AgedReceivable)
class AgedReceivableAdmin(admin.ModelAdmin):
    list_display = (
        'customer_name', 'salesperson',
        'days_1_30', 'days_31_60', 'days_61_90', 'days_91_120', 'older',
        'total', 'last_invoice_date', 'uploaded_at'
    )
    list_filter = ('salesperson', 'last_invoice_date')
    search_fields = ('customer_name', 'salesperson')
    ordering = ('-last_invoice_date',)

# Visit Data Admin Display
@admin.register(Visit)
class VisitAdmin(admin.ModelAdmin):
    list_display = (
        'visit_date',
        'sales_officer',
        'company',
        'visit_type',
        'visit_details',
        'remarks',
        'submitted_at',
    )
    list_filter = ('sales_officer', 'visit_type', 'visit_date')
    search_fields = ('sales_officer', 'company', 'visit_details', 'remarks')
    ordering = ('-visit_date',)
