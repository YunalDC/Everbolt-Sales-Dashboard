from django import template
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
from django.utils.timezone import now
from django.utils.text import slugify
from apps.home.models import Visit, Invoice, InvoiceItem
from .forms import BulkInvoiceUploadForm, BulkExcelUploadForm, TaskForm
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import models,  IntegrityError 
from django.db.models import Count, Sum, Q, F, FloatField, ExpressionWrapper
from .models import Product, CustomerOrder, Invoice, AgedReceivable, InvoiceItem
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from decimal import Decimal 
from django.conf import settings
from django.db.models.functions import Lower
from .models import UserProfile, Task, Customer
from django.db.models.functions import ExtractYear, ExtractMonth
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

import requests
import fitz  
import re
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
import calendar

# Admin Panel
@login_required
def admin_panel(request):
    return render(request, 'home/admin_panel.html')

# Dashboard
@login_required(login_url="/login/")
def index(request):
    user = request.user
    user_full_name = f"{user.first_name} {user.last_name}".strip().lower()
    invoices = Invoice.objects.filter(salesperson__iexact=user_full_name)

    total_invoices = invoices.count()
    total_amount = invoices.aggregate(Sum('total'))['total__sum'] or 0

    top_clients = (
        invoices
        .values('client')
        .annotate(order_count=Count('invoice_number'), total_revenue=Sum('total'))
        .order_by('-total_revenue')[:12]
    )

    return render(request, "home/index.html", {
        'total_invoices': total_invoices,
        'total_amount': total_amount,
        'top_clients': top_clients,
        'is_salesperson': True,
    })

# Page router
@login_required(login_url="/login/")
def pages(request):
    context = {}
    try:
        load_template = request.path.split('/')[-1]
        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))

        context['segment'] = load_template
        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:
        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))

# Visit logger
@login_required(login_url="/login/")
def mark_visit(request):
    success = None
    if request.method == 'POST':
        try:
            Visit.objects.create(
                visit_date=request.POST.get('visit_date'),
                sales_officer=request.POST.get('sales_officer'),
                company=request.POST.get('company'),
                visit_type=request.POST.get('visit_type'),
                visit_details=request.POST.get('visit_details'),
                remarks=request.POST.get('remarks'),
                submitted_at=now()
            )
            success = True
        except Exception as e:
            print("Error while saving visit:", e)
            success = False
    return render(request, 'home/icons.html', {'success': success})

# Extract Data about Customers
@login_required
def upload_customers(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "❌ Please upload a valid Excel file.")
            return render(request, 'home/upload_customers.html')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            header_map = {header.strip().lower(): idx for idx, header in enumerate(headers)}

            required_columns = ['display name', 'city', 'country', 'email', 'phone', 'salesperson']
            if not all(col in header_map for col in required_columns):
                messages.error(request, "❌ Missing required columns.")
                return render(request, 'home/upload_customers.html')

            for row in sheet.iter_rows(min_row=2, values_only=True):
                Customer.objects.create(
                    display_name=row[header_map['display name']],
                    city=row[header_map['city']],
                    country=row[header_map['country']],
                    email=row[header_map['email']],
                    phone=str(row[header_map['phone']]),
                    salesperson=row[header_map['salesperson']],
                )

            messages.success(request, "✅ Customers uploaded successfully!")

        except Exception as e:
            messages.error(request, f"❌ Error reading Excel file: {e}")

    return render(request, 'home/upload_customers.html')

# Visit Company Display

@login_required
def autocomplete_company(request):
    query = request.GET.get('q', '')
    companies = Customer.objects.filter(display_name__icontains=query).values_list('display_name', flat=True).distinct()[:10]
    return JsonResponse(list(companies), safe=False)

# Extract invoice data from PDF text (optional utility)
def extract_invoice_data(text):
    def extract(pattern, default=None, transform=lambda x: x):
        match = re.search(pattern, text, re.MULTILINE)
        return transform(match.group(1)) if match else default

    lines = [line.strip() for line in text.splitlines() if line.strip()]

    client = extract(r'Client[:\s]+([A-Za-z0-9 &().,-]+)', default="UNKNOWN")
    salesperson = "UNKNOWN"
    for i in range(min(len(lines) - 1, 20)):
        line1 = lines[i]
        line2 = lines[i + 1]
        if line1.isupper() or line2.isupper():
            continue
        if (
            all(c.isalpha() or c.isspace() for c in line1)
            and all(c.isalpha() or c.isspace() for c in line2)
        ):
            salesperson = f"{line1.strip()} {line2.strip()}"
            break

    invoice_date = extract(r'(\d{2}/\d{2}/\d{2})\s+(?:\d+\s+Days|Cash)', None, lambda x: datetime.strptime(x, "%d/%m/%y").date())
    due_date = extract(r'VILTS-\d+\s+(\d{2}/\d{2}/\d{2})', None, lambda x: datetime.strptime(x, "%d/%m/%y").date())

    return {
        "invoice_number": extract(r'(VILTS-\d+)', "UNKNOWN"),
        "client": client,
        "invoice_date": invoice_date,
        "due_date": due_date,
        "subtotal": extract(r'Sub\s+Total\s+Rs\.\s*([\d,]+\.\d{2})', 0, lambda x: float(x.replace(",", ""))),
        "vat": extract(r'VAT\s+18%\s+Rs\.\s*([\d,]+\.\d{2})', 0, lambda x: float(x.replace(",", ""))),
        "grand_total": extract(r'GRAND\s+TOTAL\s+Rs\.\s*([\d,]+\.\d{2})', 0, lambda x: float(x.replace(",", ""))),
        "salesperson": salesperson,
        "po_number": extract(r'\b(\d{7,13})\b', "-")
    }

# Invoice upload function
@login_required(login_url="/login/")
def upload_invoices(request):
    results = []
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                print("DEBUG: Excel file loaded successfully.")  # Debugging statement

                # Extract headers and map them
                headers = [cell.value.strip().lower() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                print("DEBUG: Headers extracted: ", headers)  # Debugging statement
                header_map = {header: idx for idx, header in enumerate(headers)}

                required = [
                    'number', 'client', 'invoice/bill date', 'due date', 'salesperson',
                    'total', 'product', 'quantity', 'brand', 'part number', 'unit price', 'subtotal'
                ]

                missing_cols = [col for col in required if col not in header_map]
                if missing_cols:
                    results.append(f"❌ Missing required columns: {', '.join(missing_cols)}")
                    print(f"DEBUG: Missing columns: {missing_cols}")  # Debugging statement
                else:
                    previous_invoice = None
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        number = row[header_map['number']]
                        if number:
                            try:
                                # Get general invoice info
                                invoice_number = str(number)
                                invoice_date = row[header_map['invoice/bill date']]
                                due_date = row[header_map['due date']]
                                salesperson = str(row[header_map['salesperson']])
                                client = str(row[header_map['client']])
                                total = float(str(row[header_map['total']]).replace(",", ""))
                                subtotal = float(str(row[header_map['subtotal']]).replace(",", ""))

                                print(f"DEBUG: Creating invoice {invoice_number} for client {client}")  # Debugging statement

                                # Create Invoice record
                                previous_invoice = Invoice.objects.create(
                                    invoice_number=invoice_number,
                                    client=client,
                                    invoice_date=invoice_date,
                                    due_date=due_date,
                                    salesperson=salesperson,
                                    total=total,
                                    subtotal=subtotal
                                )
                                print(f"DEBUG: Invoice {invoice_number} created successfully.")  # Debugging statement
                            except Exception as e:
                                results.append(f"❌ Error creating invoice {invoice_number}: {e}")
                                print(f"ERROR: {e}")  # Debugging statement
                                previous_invoice = None
                                continue

                        if not previous_invoice:
                            continue

                        # Now create InvoiceItems for each product line under the invoice
                        try:
                            product_name = row[header_map['product']]
                            quantity = row[header_map['quantity']]
                            brand = row[header_map['brand']]
                            part_number = row[header_map['part number']]
                            unit_price = row[header_map['unit price']]  # Unit price directly from Excel
                            line_total = row[header_map['subtotal']]  # You can calculate it if missing

                            print(f"DEBUG: Creating item for product {product_name}, quantity {quantity}")  # Debugging statement

                            # Skip creating item if product_name or quantity is missing
                            if not product_name or quantity is None:
                                raise ValueError("Missing product or quantity")

                            # Create InvoiceItem using data from Excel
                            InvoiceItem.objects.create(
                                invoice=previous_invoice,  # Link this item to the invoice
                                product_name=product_name,  # Store the product name directly
                                brand=brand,
                                part_number=part_number,
                                unit_price=unit_price,
                                quantity=quantity,
                                line_total=unit_price * quantity  # Calculate the line total
                            )
                            print(f"DEBUG: Item for product {product_name} created successfully.")  # Debugging statement
                        except Exception as e:
                            results.append(f"❌ Error creating item for invoice {previous_invoice.invoice_number}: {e}")
                            print(f"ERROR: {e}")  # Debugging statement

                    results.append("✅ Invoices and items uploaded successfully.")
            except Exception as e:
                results.append(f"❌ Failed to process Excel file: {e}")
                print(f"ERROR: Failed to process Excel file: {e}")  # Debugging statement

    return render(request, 'home/upload_invoice.html', {
        'results': results,
        'user_can_upload': request.user.is_superuser or request.user.groups.filter(name="Invoice Uploaders").exists()
    })

#Display Invoices
@login_required(login_url="/login/")
def display_invoices(request):
    query = request.GET.get('q', '')  # Get the search query from the request
    
    # Filter invoices based on the search query (invoice number or client name)
    invoices = Invoice.objects.filter(
        invoice_number__icontains=query  # This can be adjusted depending on the search field
    ).order_by('-invoice_date')  # Order invoices by the date (descending)

    # Paginate the invoices
    paginator = Paginator(invoices, 10)  # 10 invoices per page
    page_number = request.GET.get('page')  # Get the page number from the query string
    page_obj = paginator.get_page(page_number)  # Get the page object

    print("DEBUG INVOICES COUNT:", invoices.count())  # This should print the filtered count

    # Render the response with the page object and search query
    return render(request, 'home/map.html', {
        'page_obj': page_obj,  # Send the paginated invoices to the template
        'query': query  # Send the search query to retain it in the search field
    })

#Product Display
@login_required(login_url="/login/")
def display_products(request): 
    query = request.GET.get("q", "")
    product_list = Product.objects.all()

    if query:
        product_list = product_list.filter(
            Q(part_number__icontains=query) | Q(name__icontains=query)
        )

    paginator = Paginator(product_list.order_by('part_number'), 100)  # 100 per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, 'home/products.html', {
        'page_obj': page_obj,
        'query': query
    })

#Upload Products
@login_required(login_url="/login/")
def upload_products(request):
    results = []

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, "❌ No file uploaded.")
            return render(request, 'home/upload_product.html', {'results': results})

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            headers = [cell.value for cell in sheet[1]]
            header_map = {str(header).strip().lower(): idx for idx, header in enumerate(headers)}

            required = ['part number', 'name', 'brand', 'quantity on hand', 'unit of measure', 'sales price']
            if not all(col in header_map for col in required):
                messages.error(request, "❌ Excel is missing required columns.")
                return render(request, 'home/upload_product.html', {'results': results})

            from .models import Product

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    Product.objects.update_or_create(
                        part_number=row[header_map['part number']],
                        defaults={
                            'name': row[header_map['name']],
                            'brand': row[header_map['brand']],
                            'quantity_on_hand': int(row[header_map['quantity on hand']] or 0),
                            'unit_of_measure': row[header_map['unit of measure']],
                            'sales_price': float(str(row[header_map['sales price']]).replace(",", "")),
                        }
                    )
                except Exception as e:
                    results.append(f"❌ Row error: {e}")

            messages.success(request, "✅ Products uploaded successfully!")

        except Exception as e:
            messages.error(request, f"❌ Error reading Excel file: {e}")

    return render(request, 'home/upload_product.html', {'results': results})

# Mark Visit 
@csrf_exempt
@login_required(login_url="/login/")
def mark_visit(request):
    success = None
    if request.method == 'POST':
        try:
            # Get form data
            visit_date = request.POST.get('visit_date')
            sales_officer = request.POST.get('sales_officer')
            company = request.POST.get('company')
            visit_type = request.POST.get('visit_type')
            visit_details = request.POST.get('visit_details')
            remarks = request.POST.get('remarks')

            # Save to Django DB
            Visit.objects.create(
                visit_date=visit_date,
                sales_officer=sales_officer,
                company=company,
                visit_type=visit_type,
                visit_details=visit_details,
                remarks=remarks,
                submitted_at=now()
            )

            # Prepare data for Google Script
            payload = {
                'visit_date': visit_date,
                'sales_officer': sales_officer,
                'company': company,
                'visit_type': visit_type,
                'visit_details': visit_details,
                'remarks': remarks
            }

            # Send to Google Apps Script
            script_url = 'https://script.google.com/macros/s/AKfycbziaKZFcxY2Y7rqEBNBCgCc_EXDWN6ZZtJXk8lz7e5D8cgmUqz0qxl90BG411lf5fOE/exec'
            response = requests.post(script_url, data=payload, timeout=5)

            # Optional logging
            print(f"[INFO] Google Sheet response: {response.status_code} - {response.text}")

            success = response.status_code == 200

        except Exception as e:
            print("Error while saving visit:", e)
            success = False

    return render(request, 'home/icons.html', {'success': success})

# Notifications
@login_required(login_url="/login/")
def notifications_view(request):
    return render(request, "home/notifications.html")

# Assign tasks
@login_required(login_url='/login/')
def assign_tasks_view(request):
    users = User.objects.all()
    success = None
    error = None

    if request.method == "POST":
        user_id = request.POST.get("assigned_to")
        title = request.POST.get("title")
        description = request.POST.get("description")

        if user_id and title:
            try:
                assigned_user = User.objects.get(id=user_id)
                Task.objects.create(
                    assigned_to=assigned_user,
                    title=title,
                    description=description or "",
                )
                success = "✅ Task assigned successfully!"
            except Exception as e:
                error = f"❌ Error assigning task: {e}"
        else:
            error = "❌ Please fill in all required fields."

    return render(request, 'home/assign_tasks.html', {
        'users': users,
        'success': success,
        'error': error,
    })


# Collection View
@login_required(login_url='/login/')
def view_collections(request):
    receivables = AgedReceivable.objects.all().order_by('-uploaded_at')
    print("RECEIVABLES COUNT:", receivables.count())  # ✅ Add this for debugging
    return render(request, 'home/typography.html', {'receivables': receivables})

# Collection upload
@login_required(login_url='/login/')
def upload_collection(request):
    results = []

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, "❌ No file uploaded.")
            return render(request, 'home/upload_collection.html')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            headers = [cell.value for cell in sheet[1]]
            header_map = {str(header).strip().lower(): idx for idx, header in enumerate(headers)}

            required_cols = ['customer', 'sales person', '1-30', '31-60', '61-90', '91-120', 'older', 'total']
            if not all(col in header_map for col in required_cols):
                messages.error(request, "❌ Excel is missing required columns.")
                return render(request, 'home/upload_collection.html')

            AgedReceivable.objects.all().delete()  # Optional: Clear old data

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    customer_name = row[header_map['customer']]
                    salesperson = row[header_map['sales person']]

                    days_1_30 = row[header_map['1-30']] or 0
                    days_31_60 = row[header_map['31-60']] or 0
                    days_61_90 = row[header_map['61-90']] or 0
                    days_91_120 = row[header_map['91-120']] or 0
                    older = row[header_map['older']] or 0
                    total = row[header_map['total']] or 0

                    # Optional date field
                    last_invoice_date = None
                    if 'invoice date' in header_map:
                        date_val = row[header_map['invoice date']]
                        if isinstance(date_val, datetime):
                            last_invoice_date = date_val.date()

                    AgedReceivable.objects.create(
                        customer_name=customer_name,
                        salesperson=salesperson,
                        days_1_30=days_1_30,
                        days_31_60=days_31_60,
                        days_61_90=days_61_90,
                        days_91_120=days_91_120,
                        older=older,
                        total=total,
                        last_invoice_date=last_invoice_date
                    )

                except Exception as e:
                    results.append(f"❌ Failed to insert row: {e}")

            messages.success(request, "✅ Aged receivables uploaded successfully.")

        except Exception as e:
            messages.error(request, f"❌ Error reading Excel file: {e}")

    return render(request, 'home/upload_collection.html', {'results': results})

#Main Dashboard Display
@login_required(login_url="/login/")
def index(request):
    user = request.user
    full_name = f"{user.first_name} {user.last_name}".strip().lower()

    # ✅ Invoices for this user
    invoices = Invoice.objects.filter(salesperson__iexact=full_name)
    invoice_numbers = invoices.values_list('invoice_number', flat=True)

    # 🌟 Top Clients
    top_clients = (
        invoices
        .values('client')
        .annotate(order_count=Count('invoice_number'), total_revenue=Sum('total'))
        .order_by('-total_revenue')[:12]
    )

    # 🌟 Brand Sales
    tracked_brands = [
        'Schneider', 'Mennekes', 'Genebre', 'Baumer', 'Selec', 'Pilz', 'Hanyoung Nux', 'Emas',
        'SKP', 'HPC', 'Trumen', 'ONKA', 'Foxtam', 'Hensel', 'DKM', 'Perry'
    ]

    items = (
        InvoiceItem.objects
        .filter(invoice__invoice_number__in=invoice_numbers)
        .annotate(lower_brand=Lower('brand'))
        .filter(lower_brand__in=[b.lower() for b in tracked_brands])
        .annotate(
            brand_revenue=ExpressionWrapper(
                F('unit_price') * F('quantity'),
                output_field=FloatField()
            )
        )
        .values('lower_brand')
        .annotate(total_sales=Sum('brand_revenue'))
    )

    brand_data_map = {entry['lower_brand']: float(entry['total_sales']) for entry in items}
    brand_labels = tracked_brands
    brand_data = [brand_data_map.get(b.lower(), 0.0) for b in tracked_brands]

    # 🌟 Monthly Sales by Year
    monthly_sales = (
        invoices
        .annotate(year=ExtractYear('invoice_date'), month=ExtractMonth('invoice_date'))
        .values('year', 'month')
        .annotate(total=Sum('total'))
        .order_by('year', 'month')
    )

    sales_per_year = {
        2025: [0] * 12,
        2024: [0] * 12,
        2023: [0] * 12,
    }

    for entry in monthly_sales:
        year = entry['year']
        month = entry['month']
        total = entry['total']
        if year in sales_per_year and 1 <= month <= 12:
            sales_per_year[year][month - 1] = float(total)

    # 📅 Tasks assigned to this user
    user_tasks = Task.objects.filter(assigned_to=user).order_by('-created_at')
    print(f"[DEBUG] {user.username} has {user_tasks.count()} tasks.")

    return render(request, "home/index.html", {
        'total_invoices': invoices.count(),
        'total_amount': float(invoices.aggregate(Sum('total'))['total__sum'] or 0),
        'top_clients': top_clients,
        'brand_labels': json.dumps(brand_labels),
        'brand_data': json.dumps(brand_data),
        'sales_per_year': json.dumps(sales_per_year),
        'user_tasks': user_tasks,
        'is_salesperson': True,
    })

@login_required
def toggle_task_complete(request, task_id):
    if request.method == "POST":
        task = get_object_or_404(Task, id=task_id, assigned_to=request.user)
        task.is_completed = not task.is_completed
        task.save()
    return redirect("index")

@login_required
def delete_task(request, task_id):
    if request.method == "POST":
        task = get_object_or_404(Task, id=task_id, assigned_to=request.user)
        task.delete()
    return redirect("index")

# Admin Panel Display
@login_required(login_url="/login/")
def admin_panel(request):
    user = request.user

    # Allow only superusers or selected users to see sales table
    show_brand_sales = user.is_superuser or user.groups.filter(name='SalesManagers').exists()

    brand_sales_data = []

    if show_brand_sales:
        tracked_brands = [
            'Schneider', 'Mennekes', 'Genebre', 'Baumer', 'Selec', 'Pilz', 'Hanyoung Nux', 'Emas',
            'SKP', 'HPC', 'Trumen', 'ONKA', 'Foxtam', 'Hensel', 'DKM', 'Perry'
        ]

        all_users = User.objects.all()

        for salesperson in all_users:
            full_name = f"{salesperson.first_name} {salesperson.last_name}".strip().lower()
            invoices = Invoice.objects.filter(salesperson__iexact=full_name)

            invoice_numbers = invoices.values_list('invoice_number', flat=True)

            items = (
                InvoiceItem.objects
                .filter(invoice__invoice_number__in=invoice_numbers)
                .annotate(lower_brand=Lower('brand'))
                .filter(lower_brand__in=[b.lower() for b in tracked_brands])
                .annotate(
                    brand_revenue=ExpressionWrapper(
                        F('unit_price') * F('quantity'),
                        output_field=FloatField()
                    )
                )
                .values('lower_brand')
                .annotate(total_sales=Sum('brand_revenue'))
            )

            brand_map = {entry['lower_brand']: float(entry['total_sales']) for entry in items}
            brand_sales = {brand: brand_map.get(brand.lower(), 0.0) for brand in tracked_brands}

            brand_sales_data.append({
                'salesperson': salesperson.get_full_name(),
                'brand_sales': brand_sales,
            })

    return render(request, "home/admin_panel.html", {
        'show_brand_sales': show_brand_sales,
        'brand_sales_data': brand_sales_data,
    })

# Helper: only allow admin users
@login_required
def edit_users(request):
    users = User.objects.filter(is_superuser=False).prefetch_related('userprofile')

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        user = User.objects.get(id=user_id)

        if 'save' in request.POST:
            user.username = request.POST.get('username')
            user.email = request.POST.get('email')
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.save()

        elif 'upload' in request.POST:
            print("Upload image triggered ✅")
            profile_image = request.FILES.get('profile_image')
            if profile_image:
                profile, created = UserProfile.objects.get_or_create(user=user)
                profile.profile_image = profile_image
                profile.save()

        # Refresh the user object before rendering
        return redirect('edit_users')

    # Make sure to fetch latest user profile for all users
    return render(request, 'admin/edit_users.html', {'users': users})

